from flask import Flask, request, render_template, send_from_directory
from openpyxl import load_workbook
from numbers import Number
import os
import tempfile

from transcript_generator import generate_transcripts
from xlsxgeneratorV2 import split_excel_transcripts_anysheet  # Your script that splits the Excel file

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32 MB, adjust the value as needed
def get_simple_type(value):
    if isinstance(value, str):
        return "str"
    elif isinstance(value, Number):
        return "number"
    elif value is None:
        return "none"
    else:
        return "other"

def validate_excel(file_path):
    print("Validating Excel file at:", file_path)
    wb = load_workbook(file_path, data_only=False)
    issues = {}
    # Define columns that must be numeric (adjust as needed)
    numeric_columns = ["Grade", "Attendance", "Student No"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        problems = []

        headers = [cell.value for cell in ws[1]]
        unnamed = [i+1 for i, h in enumerate(headers) if h is None or str(h).strip() == ""]
        if unnamed:
            problems.append(f"Unnamed columns at positions: {unnamed}")

        header_set = set()
        repeated = []
        for h in headers:
            if h in header_set and h not in repeated and h is not None:
                repeated.append(h)
            header_set.add(h)
        if repeated:
            problems.append(f"Repeated column names: {repeated}")

        type_mismatches = {}
        for col_idx, header in enumerate(headers):
            if header is None:
                continue
            expected_type = None
            for row in ws.iter_rows(min_row=2, min_col=col_idx+1, max_col=col_idx+1):
                val = row[0].value
                if val is None:
                    continue
                val_type = get_simple_type(val)
                if expected_type is None:
                    expected_type = val_type
                elif val_type != expected_type:
                    type_mismatches.setdefault(header, []).append(val)
        if type_mismatches:
            msg = "Type mismatches found in columns:<br>"
            for k, v in type_mismatches.items():
                msg += f" - {k}: examples {v[:3]}<br>"
            problems.append(msg.strip())

        empty_columns = []
        for col_idx, header in enumerate(headers):
            if header is None:
                continue
            has_data = False
            for cell in ws.iter_cols(min_col=col_idx+1, max_col=col_idx+1, min_row=2):
                for c in cell:
                    if c.value not in [None, ""] or c.fill.fgColor.type != 'indexed' or c.data_type == "f":
                        has_data = True
                        break
                if has_data:
                    break
            if not has_data:
                empty_columns.append(header)
        if empty_columns:
            problems.append(f"Empty columns with no values: {empty_columns}")

        # Validate numeric columns
        for col_idx, header in enumerate(headers):
            if header is None:
                continue
            if header in numeric_columns:
                for row in ws.iter_rows(min_row=2, min_col=col_idx+1, max_col=col_idx+1):
                    val = row[0].value
                    if val is not None and not isinstance(val, (int, float)):
                        problems.append(f"Non-numeric value '{val}' found in numeric column '{header}' (row {row[0].row})")

        # Validate type mismatches (optional, for columns not in numeric_columns)
        type_mismatches = {}
        for col_idx, header in enumerate(headers):
            if header is None or header in numeric_columns:
                continue
            expected_type = None
            for row in ws.iter_rows(min_row=2, min_col=col_idx+1, max_col=col_idx+1):
                val = row[0].value
                if val is None:
                    continue
                val_type = get_simple_type(val)
                if expected_type is None:
                    expected_type = val_type
                elif val_type != expected_type:
                    type_mismatches.setdefault(header, []).append(val)
        if type_mismatches:
            msg = "Type mismatches found in columns:<br>"
            for k, v in type_mismatches.items():
                msg += f" - {k}: examples {v[:3]}<br>"
            problems.append(msg.strip())

        empty_cells = []
        merged_ranges = [r for r in ws.merged_cells.ranges]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for col_idx, cell in enumerate(row):
                is_merged = any(cell.coordinate in rng for rng in merged_ranges)
                is_colored = cell.fill.fgColor.type != "indexed"
                is_formula = cell.data_type == "f"

                if (cell.value is None or str(cell.value).strip() == "") and not (is_merged or is_colored or is_formula):
                    col_name = headers[col_idx] if col_idx < len(headers) else f"Col{col_idx+1}"
                    empty_cells.append(f"{col_name} (row {row_idx})")
        if empty_cells:
            problems.append(f"Empty cells found (up to 10 shown): {empty_cells[:10]}")

        if problems:
            issues[sheet_name] = problems
    return issues

@app.route("/", methods=["GET", "POST"])
def index():
    files = []
    errors = None
    if request.method == "POST":
        file = request.files.get("file")
        if file and file.filename.endswith(".xlsx"):
            temp_dir = tempfile.mkdtemp()
            file_path = os.path.join(temp_dir, file.filename)
            file.save(file_path)
            print("File saved to:", file_path)

            print("Starting validation...")
            print("Validation passed. Proceeding to split and generate transcripts.")
            try:
                split_dir = os.path.join(tempfile.gettempdir(), "split_excels")
                os.makedirs(split_dir, exist_ok=True)
                print("About to split to ", split_dir)
                small_xlsx_files, students= split_excel_transcripts_anysheet(file_path, split_dir)

                template_path = "template.docx"
                output_dir = os.path.join(tempfile.gettempdir(), "word_output")
                os.makedirs(output_dir, exist_ok=True)

                for small_file in small_xlsx_files:
                    generate_transcripts(small_file, output_dir, template_path,students)
            
                files = os.listdir(output_dir)
                return render_template("templateone.html", files=files, errors=None)
            except Exception as e:
                return f"<h3>Processing error:</h3><pre>{e}</pre>"
        return "<h3>Invalid file. Please upload a valid .xlsx file.</h3>"

    return render_template("templateone.html")

@app.route("/download/<filename>")
def download_file(filename):
    word_output_dir = os.path.join(tempfile.gettempdir(), "word_output")
    return send_from_directory(word_output_dir, filename, as_attachment=True)

if __name__ == "__main__":
    app.run(port=5000)
