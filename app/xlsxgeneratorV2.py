import os
from openpyxl import load_workbook, Workbook
from copy import copy

def split_excel_transcripts_anysheet(input_file, output_dir):
    print("Loading workbook from:", input_file)
    os.makedirs(output_dir, exist_ok=True)
    print("Output directory created:", output_dir)
    wb = load_workbook(input_file)
    # Buscar la primera hoja que tenga datos (ignorando vacías)
    print("The workbook has the following sheets:", wb.worksheets)
    raw_sheet = wb["Raw"]
    print("Using raw sheet:", raw_sheet.title)
    print("Headers in raw sheet:", [cell.value for cell in raw_sheet[1]])
    headers = [str(cell.value).strip() if cell.value else "" for cell in raw_sheet[1]]
    print("Headers processed:", headers)
    students = []
    for row in raw_sheet.iter_rows(min_row=2):
        if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
            break
        row_values = [cell.value for cell in row]
        student_dict = dict(zip(headers, row_values))
        students.append(student_dict)
    print("Number of students found:", len(students))
    print("Students data sample:", students[:2] if students else "No students found")
    print("Number of students found:", len(students))
    print("Students data sample:", students[:2] if students else "No students found")
    form_templates = {sheet: wb[sheet] for sheet in wb.sheetnames if wb[sheet] != raw_sheet}
    output_books = {}

    for student in students:
        print("Processing student:", student.get("Student No"))
        intake = student.get("Intake")
        student_no = str(student.get("Student No")).strip()
        if not intake or not student_no or intake not in form_templates:
            continue

        template = form_templates[intake]
        if intake not in output_books:
            new_wb = Workbook()
            new_wb.remove(new_wb.active)
            output_books[intake] = new_wb
        new_wb = output_books[intake]
        new_ws = new_wb.create_sheet(title=student_no)

        for row in template.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        for merged_range in template.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))
        for col_letter, col_dim in template.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width
        for row_idx, row_dim in template.row_dimensions.items():
            new_ws.row_dimensions[row_idx].height = row_dim.height

        attendance_values = []
        grade_values = []

        for key, value in student.items():
            if isinstance(key, str) and (key.lower().endswith("attendance") or key.lower().endswith("grade")):
                try:
                    parts = key.rsplit(" ", 1)
                    if len(parts) != 2:
                        continue
                    course_code, field = parts
                    field = field.lower()

                    if isinstance(value, (int, float)):
                        if field == "attendance":
                            attendance_values.append(float(value))
                        elif field == "grade":
                            grade_values.append(float(value))

                    for row in range(1, new_ws.max_row + 1):
                        if new_ws.cell(row=row, column=1).value == course_code:
                            col = 5 if field == "attendance" else 6
                            new_ws.cell(row=row, column=col, value=value)
                except:
                    continue

        avg_attendance = round(sum(attendance_values) / len(attendance_values), 2) if attendance_values else None
        avg_grade = round(sum(grade_values) / len(grade_values), 2) if grade_values else None
        print(f"Average attendance for {student_no}: {avg_attendance}, Average grade: {avg_grade}")
        if intake.startswith("BM"):
            new_ws["E19"] = avg_attendance
            new_ws["F19"] = avg_grade
        elif intake.startswith("BA"):
            new_ws["E11"] = avg_attendance
            new_ws["F11"] = avg_grade

    print("Student processing completed. Saving output files...")
    output_files = []
    for intake, wb in output_books.items():
        out_path = os.path.join(output_dir, f"{intake}_Transcripts.xlsx")
        wb.save(out_path)
        output_files.append(out_path)

    return output_files, students
