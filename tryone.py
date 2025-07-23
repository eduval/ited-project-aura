from openpyxl import load_workbook

# Ask the user to enter the file name or path
file_path = input("Enter the path to the .xlsx file: ")

try:
    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb.active  # Get the active sheet

    # Read all data
    for row in ws.iter_rows(values_only=True):
        print(row)

except FileNotFoundError:
    print("The file was not found. Please check the path.")
except Exception as e:
    print(f"An error occurred: {e}")
