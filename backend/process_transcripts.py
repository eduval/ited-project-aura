import os
import openpyxl
from openpyxl import load_workbook
from copy import copy

def process_transcripts(input_file, output_dir="transcripts_output"):
    if not input_file.endswith(".xlsx"):
        raise ValueError("Only .xlsx Excel files are supported.")

    if not os.path.exists(input_file):
        raise FileNotFoundError(f"File not found: {input_file}")

    os.makedirs(output_dir, exist_ok=True)

    wb = load_workbook(input_file)
    raw_sheet = wb["Raw"]
    headers = [str(cell.value).strip() if cell.value else "" for cell in raw_sheet[1]]
    students = [
        dict(zip(headers, [cell.value for cell in row]))
        for row in raw_sheet.iter_rows(min_row=2)
    ]
    form_templates = {sheet: wb[sheet] for sheet in wb.sheetnames if sheet != "Raw"}
    output_books = {}

    for student in students:
        intake = student.get("Intake")
        student_no = str(student.get("Student No")).strip()
        if not intake or not student_no or intake not in form_templates:
            continue

        template = form_templates[intake]
        if intake not in output_books:
            new_wb = openpyxl.Workbook()
            default_sheet = new_wb.active
            new_wb.remove(default_sheet)
            output_books[intake] = new_wb
        else:
            new_wb = output_books[intake]

        new_ws = new_wb.create_sheet(title=student_no)

        for row in template.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        for merged_range in template.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))
        for col_letter, col_dim in template.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width
        for row_idx, row_dim in template.row_dimensions.items():
            new_ws.row_dimensions[row_idx].height = row_dim.height

        # Calculate averages
        attendance_values = []
        grade_values = []

        for key, value in student.items():
            if isinstance(key, str) and (key.lower().endswith("attendance") or key.lower().endswith("grade")):
                try:
                    parts = key.rsplit(" ", 1)
                    if len(parts) != 2:
                        continue
                    course_code, field = parts
                    field = field.lower()

                    if isinstance(value, (int, float)):
                        if field == "attendance":
                            attendance_values.append(float(value))
                        elif field == "grade":
                            grade_values.append(float(value))

                    for row in range(1, new_ws.max_row + 1):
                        if new_ws.cell(row=row, column=1).value == course_code:
                            col = 5 if field == "attendance" else 6
                            new_ws.cell(row=row, column=col, value=value)
                except Exception as e:
                    print(f"⚠️ Error processing {key}: {e}")
                    continue

        avg_attendance = round(sum(attendance_values) / len(attendance_values), 2) if attendance_values else None
        avg_grade = round(sum(grade_values) / len(grade_values), 2) if grade_values else None

        print(f"✅ {student_no} | Intake: {intake} | Avg Attendance: {avg_attendance} | Avg Grade: {avg_grade}")

        if intake.startswith("BM"):
            new_ws["E19"] = avg_attendance
            new_ws["F19"] = avg_grade
        elif intake.startswith("BA"):
            new_ws["E11"] = avg_attendance
            new_ws["F11"] = avg_grade

    for intake, wb in output_books.items():
        out_path = os.path.join(output_dir, f"{intake}_Transcripts.xlsx")
        wb.save(out_path)

    print(f"✅ All transcripts generated in: {output_dir}")
    return output_dir
