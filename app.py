from flask import Flask, request, render_template
from openpyxl import load_workbook
import os
import tempfile

app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def upload_file():
    if request.method == "POST":
        file = request.files["file"]

        # ✅ Ensure it's an Excel file with .xlsx extension
        if file and file.filename.endswith(".xlsx"):
            # ✅ Save the uploaded file to a temporary location
            temp_dir = tempfile.mkdtemp()
            file_path = os.path.join(temp_dir, file.filename)
            file.save(file_path)

            try:
                # ✅ Call flexible processor that works with any Excel structure
                sheet_summary = process_excel(file_path)

                # ✅ Build HTML response: sheet names + sample rows
                html_output = "<h2>Uploaded Excel File Processed</h2>"
                for sheet, rows in sheet_summary.items():
                    html_output += f"<h3>Sheet: {sheet}</h3><pre>"
                    for row in rows[:10]:  # Only show first 10 rows to keep output readable
                        html_output += str(row) + "\n"
                    html_output += "</pre>"

                return html_output

            except Exception as e:
                return f"<h3>Error while processing Excel file:</h3><pre>{e}</pre>"

        return "Invalid file type. Please upload a .xlsx file."

    # ✅ On GET, show the upload form
    return render_template("upload.html")


# ✅ This is the new universal processor for any Excel file
def process_excel(input_file):
    """
    Opens any .xlsx file, reads all sheets and returns their data.
    """
    wb = load_workbook(input_file, data_only=True)
    sheet_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []

        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))  # Convert tuple to list for flexibility

        sheet_data[sheet_name] = rows

    return sheet_data
